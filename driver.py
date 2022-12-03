import openpyxl
from selenium import webdriver
import time

#from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc
import accountInfo

def prepare_form(originsheet, tempsheet):
    global NUM_OF_ROWS
    NUM_OF_ROWS = -1

    # 이메일 셀 범위 복사 붙여넣기 (C열-> L열로 복사)
    columns = originsheet.iter_cols(min_col=3, max_col=3)
    for col in columns:
        for cell in col:
            cell_new = originsheet.cell(row=cell.row, column=12, value=cell.value)
            NUM_OF_ROWS += 1  # 총 인원 수는 (NUM_OF_ROWS-1) 명

    # 참조용 메일 주소 넣기
    originsheet.cell(row=NUM_OF_ROWS + 2, column=12).value = "dl_edu_cert@navercorp.com"
    originsheet.cell(row=NUM_OF_ROWS + 3, column=12).value = "dl_certification@navercorp.com"

    # 날짜 삽입 (A열-> M열)
    dateValue = (originsheet.cell(row=2, column=1).value) % 10000  ## 1102
    month = (int)(dateValue / 100)
    day = (dateValue % 100)
    dateString = str(month) + "월 " + str(day) + "일"  # 월-일 형태의 스트링으로 변환
    for i in range(2, NUM_OF_ROWS + 4):
        originsheet.cell(row=i, column=13).value = dateString  # M열에 삽입
    print("--FORM GENERATED--")

    # tempsheet의 (2,1) ~ (NUM_OF_ROWS+3, 2) 범위에 originsheet의 (2, 12) ~ (NUM_OF_ROWS+3, 13) 범위의 값을 copy&paste
    # tempsheet를 data에 저장--> wb.save("temp_today.xlsx")
    print("--FORM PREPARED--")
def set_mailer_form(mailsheet):
    # drop_duplicate
    global NUM_OF_PEOPLE
    # NUM_OF_PEOPLE =

def login(driver):  ### 초기 1회 ###
    # '로그인' 클릭
    driver.find_element(By.CSS_SELECTOR, "a.glue-header__link").click()
    time.sleep(2)
    # 아이디 입력 & '다음' 클릭
    driver.find_element(By.CSS_SELECTOR, "input#identifierId").send_keys(accountInfo.admin_id)
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#identifierNext > div > button > span").click()
    time.sleep(2)
    # 비밀번호 입력 & '다음' 클릭
    driver.find_element(By.CSS_SELECTOR, "#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input").send_keys(
        accountInfo.admin_pw)
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#passwordNext > div > button > span").click()
    time.sleep(5)

    ### 전화번호 인증창이 나타날 경우
    try:
        phoneNumInput = driver.find_element(By.CSS_SELECTOR, "#phoneNumberId")
        myPhoneNum = input("인증번호를 받을 전화번호를 입력하세요: ")
        phoneNumInput.send_keys(myPhoneNum)
        time.sleep(2)
        ## '다음' 버튼: #idvanyphonecollectNext > div > button > span
        driver.find_element(By.CSS_SELECTOR, "#idvanyphonecollectNext > div > button > span").click()
        time.sleep(1)
        authNumInput = driver.find_element(By.CSS_SELECTOR, "#idvAnyPhonePin")
        myAuthNum = input("문자메시지로 받은 인증번호 6자리를 입력하세요: ")
        authNumInput.send_keys(myAuthNum)
        time.sleep(2)
        ## '다음' 버튼: #idvanyphoneverifyNext > div > button > div.VfPpkd-RLmnJb
        driver.find_element(By.CSS_SELECTOR, "# idvanyphoneverifyNext > div > button > div.VfPpkd-RLmnJb").click()
        print("--Your Phone Number is validated--")
    except: #NoSuchElementException
        print("--Authentification Success--")
    print("--LOGIN COMPLETED--")
def generate_link(driver, mailsheet, current_row):
    # '새 회의' 버튼
    driver.find_element(By.CSS_SELECTOR,
                        "#yDmH0d > c-wiz > div > div.S3RDod > div > div.Qcuypc > div.Ez8Iud > div > div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(1) > div > button > span").click()
    time.sleep(1)
    # '나중에 진행할 회의 만들기' 버튼
    driver.find_element(By.CSS_SELECTOR,
                        "#yDmH0d > c-wiz > div > div.S3RDod > div > div.Qcuypc > div.Ez8Iud > div > div.VfPpkd-xl07Ob-XxIAqe-OWXEXe-oYxtQd > div:nth-child(2) > div > ul > li:nth-child(2) > span.VfPpkd-StrnGf-rymPhb-b9t22c").click()
    time.sleep(5)
    # 링크 값 가져오기
    try:
        link = driver.find_element(By.CSS_SELECTOR, "#yDmH0d > div.VfPpkd-Sx9Kwc.VfPpkd-Sx9Kwc-OWXEXe-vOE8Lb.cC1eCc.UDxLd.PzCPDd.VKf0Le.u9lF8e.VfPpkd-Sx9Kwc-OWXEXe-FNFY6c > div.VfPpkd-wzTsW > div > div.VfPpkd-cnG4Wd > div > div:nth-child(2) > div > div.NgL38b.CZ8zsc > div.VA2JSc")
        #print("USING OLD CHROME VERSION..")
    except:
        link = driver.find_element(By.CSS_SELECTOR, "#yDmH0d > div.VfPpkd-Sx9Kwc.VfPpkd-Sx9Kwc-OWXEXe-vOE8Lb.cC1eCc.UDxLd.PzCPDd.VKf0Le.u9lF8e.VfPpkd-Sx9Kwc-OWXEXe-FNFY6c > div.VfPpkd-wzTsW > div > div.VfPpkd-cnG4Wd > div > div:nth-child(2) > div > div.Hayy8b")
        #print("USING LATEST CHROME VERSION..")
    print(link.text)

    # 셀에 삽입
    #sheet.cell(row=current_row, column=14).value = link.text
    mailsheet.cell(row=current_row, column=3).value = link.text
    # 'X' 버튼
    driver.find_element(By.CSS_SELECTOR,
                        "div.u9lF8e div.VfPpkd-oclYLd > button > span > svg").click()
    time.sleep(1)

def set_result_form(mailsheet, tempsheet):
    # 1) mailsheet의 (1,1) ~ (NUM_OF_PEOPLE,3)
    # mailsheet에서 (i,1)값을 가져와 resultsheet에서 찾는다
    # resultsheet에 있는 값만큼
    # NUM_OF_PEOPLE 만큼 반복
    return tempsheet

if __name__=="__main__":
    # 초기 시트
    wb = openpyxl.load_workbook("./data/자격시험수험자리스트_1101.xlsx")
    originsheet = wb.active
    # 중복 체크용 임시 시트
    wb_temp = openpyxl.load_workbook("./mailer/MASS_INPUT_FORM.xlsx")
    tempsheet = wb_temp.active

    print("START!")
    prepare_form(originsheet)

    # 메일 전송용 시트
    wb_mail = openpyxl.load_workbook("./data/temp_1101.xlsx")
    mailsheet = wb_mail.active
    set_mailer_form(mailsheet)

    driver = uc.Chrome()
    driver.get('https://meet.google.com/')
    driver.maximize_window()
    time.sleep(2)
    login(driver) #로그인
    for i in range(2, NUM_OF_PEOPLE+2):
        generate_link(driver, mailsheet, i)

    # mailersheet 생성이 완료되었다면 중복응시자를 같은값으로 체크하여 tempsheet에 링크를 넣는다
    set_result_form(mailsheet, tempsheet)
    wb_mail.save("./mailer/MASS_INPUT_FORM_1101.xlsx")

    print("END!")
    wb.save("./result/자격시험수험자리스트_1101_최종.xlsx")

"""
# 나중에) 메일 전송용 엑셀 만들기
wb_mail = openpyxl.Workbook()
mailsheet = wb_mail.active
mailsheet.append(["수신자 Email 주소", "date", "link"])  # <--맨 첫 행에 추가
wb_mail.save("./result/MASS_INPUT_FORM_today.xlsx")
"""
