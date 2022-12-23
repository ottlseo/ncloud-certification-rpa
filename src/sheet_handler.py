import openpyxl
from selenium import webdriver
import time

def prepare_form(originsheet, tempsheet):
    global NUM_OF_ROWS
    NUM_OF_ROWS = -1

    # 이메일 셀 범위 복사 붙여넣기 (C열-> L열로 복사)
    columns = originsheet.iter_cols(min_col=3, max_col=3)
    for col in columns:
        for cell in col:
            cell_new = originsheet.cell(row=cell.row, column=12, value=cell.value)
            NUM_OF_ROWS += 1  # 총 인원 수는 (NUM_OF_ROWS-1) 명

    # 참조용 메일 주소 넣기
    originsheet.cell(row=NUM_OF_ROWS + 2, column=12).value = "dl_edu_cert@navercorp.com"
    originsheet.cell(row=NUM_OF_ROWS + 3, column=12).value = "dl_certification@navercorp.com"

    # 날짜 삽입 (A열-> M열)
    dateValue = (originsheet.cell(row=2, column=1).value) % 10000  ## 1102
    month = (int)(dateValue / 100)
    day = (dateValue % 100)
    dateString = str(month) + "월 " + str(day) + "일"  # 월-일 형태의 스트링으로 변환
    for i in range(2, NUM_OF_ROWS + 4):
        originsheet.cell(row=i, column=13).value = dateString  # M열에 삽입
    print("--FORM GENERATED--")

    # tempsheet의 (2,1) ~ (NUM_OF_ROWS+3, 2) 범위에 originsheet의 (2, 12) ~ (NUM_OF_ROWS+3, 13) 범위의 값을 copy&paste
    # tempsheet를 data에 저장--> wb.save("temp_today.xlsx")
    print("--FORM PREPARED--")
def set_mailer_form(mailsheet):
    # drop_duplicate
    global NUM_OF_PEOPLE
    # NUM_OF_PEOPLE =



def set_result_form(mailsheet, tempsheet):
    # 1) mailsheet의 (1,1) ~ (NUM_OF_PEOPLE,3)
    # mailsheet에서 (i,1)값을 가져와 resultsheet에서 찾는다
    # resultsheet에 있는 값만큼
    # NUM_OF_PEOPLE 만큼 반복
    return tempsheet

if __name__=="__main__":

    # 초기 시트
    wb = openpyxl.load_workbook("../data/자격시험수험자리스트_1101.xlsx")
    originsheet = wb.active
    # 중복 체크용 임시 시트
    wb_temp = openpyxl.load_workbook("../mailer/MASS_INPUT_FORM.xlsx")
    tempsheet = wb_temp.active

    print("START!")
    prepare_form(originsheet)

    # 메일 전송용 시트
    wb_mail = openpyxl.load_workbook("../data/temp_1101.xlsx")
    mailsheet = wb_mail.active
    set_mailer_form(mailsheet)

    # mailersheet 생성이 완료되었다면 중복응시자를 같은값으로 체크하여 tempsheet에 링크를 넣는다
    set_result_form(mailsheet, tempsheet)
    wb_mail.save("../mailer/MASS_INPUT_FORM_1101.xlsx")

    print("END!")
    wb.save("../result/자격시험수험자리스트_1101_최종.xlsx")